import argparse
import html
import http.cookiejar
import json
import os
from pathlib import Path
import re
import ssl
import subprocess
import time as time_module
import unicodedata
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from collections import Counter
from datetime import date, datetime, time, timedelta, timezone

from openpyxl import load_workbook

try:
    from yt_dlp import YoutubeDL
except Exception:
    YoutubeDL = None

try:
    from keybert import KeyBERT
except Exception:
    KeyBERT = None

try:
    from sentence_transformers import SentenceTransformer
except Exception:
    SentenceTransformer = None


YT_NS = {"atom": "http://www.w3.org/2005/Atom", "yt": "http://www.youtube.com/xml/schemas/2015"}

TOPIC_LABELS = [
    "Wspolpraca i umowy",
    "Webinar lub szkolenie",
    "Wydarzenie branzowe",
    "Promocja lub konkurs",
    "Nowy produkt lub rozwiazanie",
    "Nagroda lub wyroznienie",
    "Rekrutacja i kariera",
    "Poradnik i wskazowki",
    "Aktualizacja firmowa",
]

TOPIC_OLD_GENERIC_MAP = {
    "podpisanie umowy wspolpraca": "Wspolpraca i umowy",
    "informacja o webinarze szkoleniu": "Webinar lub szkolenie",
    "informacja o webinarze / szkoleniu": "Webinar lub szkolenie",
    "informacja o wydarzeniu": "Wydarzenie branzowe",
    "promocja konkurs": "Promocja lub konkurs",
    "prezentacja produktu": "Nowy produkt lub rozwiazanie",
    "wyroznienie ranking": "Nagroda lub wyroznienie",
    "informacja rekrutacyjna": "Rekrutacja i kariera",
    "poradnik wskazowki": "Poradnik i wskazowki",
    "poradnik / wskazowki": "Poradnik i wskazowki",
    "aktualizacja firmowa": "Aktualizacja firmowa",
}

TOPIC_RULES = [
    (["podpis", "umow", "porozumien", "wspolprac", "partnerstw"], "Wspolpraca i umowy"),
    (["webinar", "szkolen", "na zywo", "live"], "Webinar lub szkolenie"),
    (["targ", "wydarzen", "event", "konferenc", "stoisk"], "Wydarzenie branzowe"),
    (["promoc", "konkurs", "rabat", "akcj"], "Promocja lub konkurs"),
    (["premier", "nowy produkt", "nowa odslona", "modul", "falownik", "inwerter", "rekuperator"], "Nowy produkt lub rozwiazanie"),
    (["nagrod", "wyrozn", "ranking", "award"], "Nagroda lub wyroznienie"),
    (["rekrut", "karier", "praca", "zatrudn"], "Rekrutacja i kariera"),
    (["poradnik", "jak ", "wskazowk", "instrukcj"], "Poradnik i wskazowki"),
]

TOPIC_DEFAULTS = {
    "Wspolpraca i umowy": "zakres partnerstwa, strony wspolpracy oraz cel wspolnego projektu",
    "Webinar lub szkolenie": "temat spotkania, grupa odbiorcow oraz najwazniejsze informacje organizacyjne",
    "Wydarzenie branzowe": "zapowiedz udzialu firmy i glowny obszar prezentowanych rozwiazan",
    "Promocja lub konkurs": "warunki akcji, czas trwania i korzysci dla odbiorcow",
    "Nowy produkt lub rozwiazanie": "najwazniejsze funkcje produktu oraz jego praktyczne zastosowania",
    "Nagroda lub wyroznienie": "powod wyroznienia i jego znaczenie dla firmy lub rynku",
    "Rekrutacja i kariera": "stanowisko, wymagania i podstawowe informacje dla kandydatow",
    "Poradnik i wskazowki": "praktyczne wskazowki, najczestsze bledy oraz rekomendowane kroki",
    "Aktualizacja firmowa": "najwazniejsze informacje dotyczace dzialan firmy i kierunku komunikacji",
}

TOPIC_STOPWORDS = {
    "a",
    "aby",
    "albo",
    "ale",
    "bo",
    "by",
    "byc",
    "byly",
    "co",
    "czy",
    "dla",
    "do",
    "dotyczy",
    "firmy",
    "i",
    "ich",
    "jak",
    "jest",
    "juz",
    "kluczowe",
    "na",
    "nad",
    "nas",
    "nasza",
    "nasze",
    "naszej",
    "nie",
    "o",
    "od",
    "oraz",
    "post",
    "posta",
    "przez",
    "sie",
    "ta",
    "ten",
    "to",
    "u",
    "w",
    "we",
    "z",
    "za",
    "about",
    "and",
    "for",
    "from",
    "in",
    "is",
    "of",
    "on",
    "our",
    "the",
    "this",
    "to",
    "with",
}

TOPIC_GENERIC_TOKENS = {
    "aktualizacja",
    "firmowa",
    "informacja",
    "post",
    "posta",
    "kluczowe",
    "watek",
    "watki",
    "promocja",
    "konkurs",
    "produkt",
    "funkcje",
    "zastosowanie",
    "poradnik",
    "wskazowki",
    "webinar",
    "szkolenie",
    "wydarzenie",
    "wspolpraca",
    "umowy",
}

_TOPIC_AI_MODEL = None
_TOPIC_AI_KW = None
_TOPIC_AI_DISABLED = False
_TOPIC_LLM_PROVIDER = None
_TOPIC_LLM_DISABLED = False
_RAW_TEXT_URL_CACHE = {}
_TOPIC_OLLAMA_MODEL = None
_META_CONFIG_CACHE = None
DEFAULT_SKIPPED_PLATFORMS = ["tiktok"]


def normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def clean_text(value, max_len=500):
    if value is None:
        return ""
    txt = html.unescape(str(value))
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt[:max_len]


def strip_tags(raw):
    return clean_text(re.sub(r"<[^>]+>", " ", raw or ""))


def normalize_raw_text_content(value):
    txt = clean_text(value, 1800)
    txt = strip_tags(txt)
    txt = re.sub(r"^\d+\s+likes?,\s*\d+\s+comments?\s*-\s*[^:]+:\s*", "", txt, flags=re.I)
    txt = re.sub(r"^\d+\s+likes?\s*-\s*[^:]+:\s*", "", txt, flags=re.I)
    txt = re.sub(r"\bSee translation\b", "", txt, flags=re.I)
    txt = txt.strip(" \"'`")
    txt = re.sub(r"\s+", " ", txt).strip()
    return clean_text(txt, 1800)


def ascii_fold(value):
    txt = clean_text(value, 2000)
    return "".join(ch for ch in unicodedata.normalize("NFKD", txt) if not unicodedata.combining(ch)).lower()


def word_tokens(text, lower=False):
    tokens = re.findall(r"[^\W_]+", clean_text(text, 2400), flags=re.UNICODE)
    if lower:
        return [t.lower() for t in tokens]
    return tokens


def summarize_topic(text, post_url=""):
    topic, _quality, _reason = summarize_topic_with_meta(text, post_url=post_url)
    return topic


def summarize_topic_with_meta(text, post_url=""):
    raw_original = normalize_raw_text_content(text)
    if not raw_original:
        return "", "low", "Brak tresci zrodlowej (raw_text)."

    raw = unwrap_existing_topic(raw_original)
    folded = normalize_for_rules(raw)
    label = classify_topic_label(folded)

    llm_context, llm_mode = build_llm_topic_context(raw, label)
    if llm_context:
        return clean_text(f"{label}: {llm_context}", 260), "high", f"Generowane przez LLM ({llm_mode})."

    ai_context = build_ai_topic_context(raw, label)
    if ai_context:
        return clean_text(f"{label}: {ai_context}", 260), "medium", "Generowane przez NLP (SentenceTransformer + KeyBERT)."

    fallback_context = build_fallback_topic_context(raw, label)
    fallback_quality = "medium" if len(word_tokens(fallback_context)) >= 9 else "low"
    fallback_reason = "Fallback z raw_text." if fallback_quality == "medium" else "Fallback: zbyt malo tresci w raw_text."
    return clean_text(f"{label}: {fallback_context}", 240), fallback_quality, fallback_reason


def build_topic_payload(raw_text, post_url=""):
    raw = normalize_raw_text_content(raw_text)
    if not raw:
        return {
            "raw_text": "",
            "topic": "",
            "topic_quality": "low",
            "topic_reason": "Brak tresci zrodlowej (raw_text).",
        }
    topic, quality, reason = summarize_topic_with_meta(raw, post_url=post_url)
    return {
        "raw_text": raw,
        "topic": topic,
        "topic_quality": quality,
        "topic_reason": reason,
    }


def unwrap_existing_topic(raw_text):
    raw = clean_text(raw_text, 1800)
    changed = True
    while changed:
        changed = False
        for lbl in TOPIC_LABELS:
            prefix = f"{lbl}:"
            if raw.lower().startswith(prefix.lower()):
                raw = clean_text(raw[len(prefix):], 1800)
                changed = True
    raw = re.sub(r"(?i)^kluczowe watki:\s*", "", raw)
    raw = re.sub(r"(?i)^najwazniejsze watki:\s*", "", raw)
    raw = re.sub(r"(?i)\b(kluczowe|najwazniejsze)\s+watki\s*:?.*$", "", raw)
    return clean_text(raw, 1800)


def normalize_for_rules(text):
    folded = ascii_fold(text)
    folded = re.sub(r"https?://\S+", " ", folded)
    folded = re.sub(r"[@#][\w\-]+", " ", folded)
    folded = re.sub(r"\s+", " ", folded).strip()
    return folded


def classify_topic_label(folded_text):
    for keywords, candidate in TOPIC_RULES:
        if any(k in folded_text for k in keywords):
            return candidate
    folded_compact = re.sub(r"[^a-z0-9 ]", " ", folded_text)
    folded_compact = re.sub(r"\s+", " ", folded_compact).strip()
    if folded_compact in TOPIC_OLD_GENERIC_MAP:
        return TOPIC_OLD_GENERIC_MAP[folded_compact]
    return "Aktualizacja firmowa"


def topic_script_dir():
    return Path(__file__).resolve().parent


def topic_project_root():
    return topic_script_dir().parent


def local_ollama_base_url():
    return clean_text(os.getenv("OLLAMA_BASE_URL", "http://127.0.0.1:11434"), 200).rstrip("/")


def local_ollama_models_dir():
    env_dir = clean_text(os.getenv("OLLAMA_MODELS", ""), 500)
    if env_dir:
        return Path(env_dir)
    return topic_project_root() / "tools" / "ollama-download" / "models"


def find_local_ollama_exe():
    candidates = []
    env_exe = clean_text(os.getenv("OLLAMA_EXE", ""), 500)
    if env_exe:
        candidates.append(Path(env_exe))
    candidates.append(topic_project_root() / "tools" / "ollama-download" / "ollama" / "ollama.exe")
    candidates.append(topic_project_root() / "tools" / "ollama" / "ollama.exe")
    for candidate in candidates:
        if candidate and candidate.exists():
            return candidate
    return None


def fetch_ollama_tags(base_url):
    try:
        req = urllib.request.Request(f"{base_url}/api/tags", headers={"Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            body = resp.read().decode("utf-8", "ignore")
        return json.loads(body)
    except Exception:
        return None


def local_ollama_is_ready(base_url):
    data = fetch_ollama_tags(base_url)
    return data if isinstance(data, dict) and "models" in data else None


def start_local_ollama_server():
    base_url = local_ollama_base_url()
    ready = local_ollama_is_ready(base_url)
    if ready is not None:
        return True

    exe_path = find_local_ollama_exe()
    if exe_path is None:
        return False

    models_dir = local_ollama_models_dir()
    models_dir.mkdir(parents=True, exist_ok=True)

    env = os.environ.copy()
    for key in ["HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "GIT_HTTP_PROXY", "GIT_HTTPS_PROXY"]:
        env[key] = ""
    env["OLLAMA_MODELS"] = str(models_dir)
    env["OLLAMA_HOST"] = base_url

    creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0) | getattr(subprocess, "DETACHED_PROCESS", 0)
    try:
        subprocess.Popen(
            [str(exe_path), "serve"],
            cwd=str(exe_path.parent),
            env=env,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=creationflags,
        )
    except Exception:
        return False

    for _ in range(20):
        time_module.sleep(1)
        ready = local_ollama_is_ready(base_url)
        if ready is not None:
            return True
    return False


def detect_local_ollama_model():
    global _TOPIC_OLLAMA_MODEL
    if _TOPIC_OLLAMA_MODEL:
        return _TOPIC_OLLAMA_MODEL

    preferred = clean_text(os.getenv("OLLAMA_MODEL", ""), 120)
    if preferred:
        _TOPIC_OLLAMA_MODEL = preferred
        return _TOPIC_OLLAMA_MODEL

    if not start_local_ollama_server():
        return ""

    data = local_ollama_is_ready(local_ollama_base_url()) or {}
    installed = [clean_text((item or {}).get("name"), 120) for item in data.get("models", [])]
    installed = [name for name in installed if name]
    if not installed:
        return ""

    preferred_names = ["qwen2.5:0.5b", "qwen2.5:1.5b", "qwen2.5:3b", "llama3.2:3b"]
    for preferred_name in preferred_names:
        if preferred_name in installed:
            _TOPIC_OLLAMA_MODEL = preferred_name
            return _TOPIC_OLLAMA_MODEL

    _TOPIC_OLLAMA_MODEL = installed[0]
    return _TOPIC_OLLAMA_MODEL


def get_topic_llm_provider():
    global _TOPIC_LLM_PROVIDER, _TOPIC_LLM_DISABLED
    if _TOPIC_LLM_DISABLED:
        return None
    if _TOPIC_LLM_PROVIDER is not None:
        return _TOPIC_LLM_PROVIDER if _TOPIC_LLM_PROVIDER != "none" else None

    if detect_local_ollama_model():
        _TOPIC_LLM_PROVIDER = "ollama"
        return _TOPIC_LLM_PROVIDER
    if clean_text(os.getenv("OPENAI_API_KEY", ""), 200):
        _TOPIC_LLM_PROVIDER = "openai"
        return _TOPIC_LLM_PROVIDER
    _TOPIC_LLM_PROVIDER = "none"
    return None


def post_json_request(url, payload, headers=None, timeout=22):
    req_headers = {"Content-Type": "application/json"}
    if headers:
        req_headers.update(headers)
    req = urllib.request.Request(
        url=url,
        data=json.dumps(payload).encode("utf-8"),
        headers=req_headers,
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        body = resp.read()
    return json.loads(body.decode("utf-8", errors="ignore"))


def clean_generated_sentence(text, max_words=30):
    out = clean_text(text, 600)
    out = re.sub(r"https?://\S+", " ", out)
    out = re.sub(r"[@#][\w\-]+", " ", out)
    out = re.sub(r"[\"'`]+", "", out)
    out = re.sub(r"\s+", " ", out).strip(" -,:;.")
    words = out.split()
    if len(words) > max_words:
        out = " ".join(words[:max_words]).strip(" -,:;.")
    return clean_text(out, 260)


def call_ollama_topic_context(raw_text, label):
    model_name = detect_local_ollama_model()
    if not model_name:
        return ""
    base_url = local_ollama_base_url()
    prompt = (
        "Napisz jedno konkretne zdanie po polsku, ktore streszcza post firmowy.\n"
        f"Kategoria: {label}\n"
        "Wymagania: bez hashtagow, bez emotikon, max 28 slow, jasno co sie wydarzylo i czego dotyczy.\n"
        f"Tresc posta:\n{raw_text}\n"
        "Wynik:"
    )
    payload = {
        "model": model_name,
        "prompt": prompt,
        "stream": False,
        "options": {"temperature": 0.2, "num_predict": 80, "num_ctx": 1024},
    }
    try:
        data = post_json_request(f"{base_url}/api/generate", payload, timeout=45)
        return clean_generated_sentence(data.get("response", ""), max_words=28)
    except Exception:
        return ""


def parse_openai_output(data):
    text = clean_text(data.get("output_text", ""), 600)
    if text:
        return text
    for item in data.get("output", []):
        for content in item.get("content", []):
            chunk = clean_text(content.get("text", ""), 600)
            if chunk:
                return chunk
    return ""


def call_openai_topic_context(raw_text, label):
    api_key = clean_text(os.getenv("OPENAI_API_KEY", ""), 300)
    if not api_key:
        return ""
    model_name = clean_text(os.getenv("OPENAI_MODEL", "gpt-4.1-mini"), 100)
    payload = {
        "model": model_name,
        "temperature": 0.2,
        "max_output_tokens": 120,
        "input": [
            {
                "role": "system",
                "content": [
                    {
                        "type": "input_text",
                        "text": "Tworzysz zwiezle opisy postow konkurencji. Zwracaj jedno zdanie po polsku, rzeczowo i konkretnie.",
                    }
                ],
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_text",
                        "text": (
                            f"Kategoria: {label}\n"
                            "Wymagania: 14-28 slow, bez hashtagow i bez ogolnikow.\n"
                            "Napisz, co sie wydarzylo i czego dotyczy komunikat.\n"
                            f"Tresc posta:\n{raw_text}"
                        ),
                    }
                ],
            },
        ],
    }
    headers = {"Authorization": f"Bearer {api_key}"}
    try:
        data = post_json_request("https://api.openai.com/v1/responses", payload, headers=headers, timeout=28)
        return clean_generated_sentence(parse_openai_output(data), max_words=28)
    except Exception:
        return ""


def build_llm_topic_context(raw_text, label):
    global _TOPIC_LLM_DISABLED
    provider = get_topic_llm_provider()
    if not provider:
        return "", ""
    out = ""
    if provider == "ollama":
        out = call_ollama_topic_context(raw_text, label)
    elif provider == "openai":
        out = call_openai_topic_context(raw_text, label)
    if out and len(word_tokens(out)) >= 8:
        return out, provider
    # Disable for current run after first failed attempt to avoid repeated slow retries.
    _TOPIC_LLM_DISABLED = True
    return "", ""


def init_topic_ai():
    global _TOPIC_AI_MODEL, _TOPIC_AI_KW, _TOPIC_AI_DISABLED
    if _TOPIC_AI_DISABLED:
        return None, None
    if _TOPIC_AI_MODEL is not None:
        return _TOPIC_AI_MODEL, _TOPIC_AI_KW
    if SentenceTransformer is None:
        return None, None

    try:
        _TOPIC_AI_MODEL = SentenceTransformer("paraphrase-multilingual-MiniLM-L12-v2", local_files_only=True)
    except Exception:
        _TOPIC_AI_DISABLED = True
        return None, None

    _TOPIC_AI_KW = None
    if KeyBERT is not None:
        try:
            _TOPIC_AI_KW = KeyBERT(model=_TOPIC_AI_MODEL)
        except Exception:
            _TOPIC_AI_KW = None
    return _TOPIC_AI_MODEL, _TOPIC_AI_KW


def cosine_similarity(a, b):
    if not a or not b:
        return 0.0
    num = sum(float(x) * float(y) for x, y in zip(a, b))
    den_a = sum(float(x) * float(x) for x in a) ** 0.5
    den_b = sum(float(y) * float(y) for y in b) ** 0.5
    den = den_a * den_b
    if den == 0:
        return 0.0
    return num / den


def split_sentences(text):
    parts = re.split(r"(?<=[\.\!\?])\s+|\n+", clean_text(text, 1800))
    out = []
    for part in parts:
        p = clean_text(part, 350).strip(" -,:;")
        if len(p) >= 24:
            out.append(p)
    return out[:8]


def looks_like_template_source(text):
    folded = normalize_for_rules(text)
    if not folded:
        return True
    if folded in TOPIC_OLD_GENERIC_MAP:
        return True
    for default_text in TOPIC_DEFAULTS.values():
        default_folded = normalize_for_rules(default_text)
        if default_folded and default_folded in folded and len(folded.split()) <= 22:
            return True
    tokens = [t for t in word_tokens(folded, lower=True) if len(t) > 2]
    if len(tokens) <= 6:
        return True
    uniq_ratio = len(set(tokens)) / max(1, len(tokens))
    return uniq_ratio < 0.45


def pick_lead_sentence(model, source_text):
    sentences = split_sentences(source_text)
    if not sentences:
        return ""
    if len(sentences) == 1 or model is None:
        return clean_text(sentences[0], 170)
    try:
        embeddings = model.encode(sentences + [source_text], convert_to_numpy=False)
        src = embeddings[-1]
        best_idx = 0
        best_score = -1.0
        for idx, vec in enumerate(embeddings[:-1]):
            score = cosine_similarity(vec, src)
            if score > best_score:
                best_score = score
                best_idx = idx
        return clean_text(sentences[best_idx], 170)
    except Exception:
        return clean_text(sentences[0], 170)


def phrase_tokens(phrase):
    return [tok for tok in word_tokens(phrase, lower=True) if tok not in TOPIC_STOPWORDS and len(tok) >= 3]


def jaccard(a, b):
    if not a or not b:
        return 0.0
    sa = set(a)
    sb = set(b)
    return len(sa.intersection(sb)) / max(1, len(sa.union(sb)))


def extract_keyphrases(kw_model, source_text):
    if kw_model is None:
        return extract_keyphrases_heuristic(source_text)
    try:
        raw_keywords = kw_model.extract_keywords(
            source_text,
            keyphrase_ngram_range=(1, 3),
            stop_words=list(TOPIC_STOPWORDS),
            use_mmr=True,
            diversity=0.7,
            top_n=10,
        )
    except Exception:
        return extract_keyphrases_heuristic(source_text)

    phrases = []
    phrase_tokens_acc = []
    seen = set()
    for phrase, score in raw_keywords:
        p = clean_text(phrase, 90).strip(" -,:;.")
        if len(p) < 4:
            continue
        folded = ascii_fold(p)
        if folded in seen or folded in TOPIC_STOPWORDS:
            continue
        token_count = len(word_tokens(p))
        if token_count > 5:
            continue
        if score is not None and float(score) < 0.15:
            continue
        ptok = phrase_tokens(p)
        if not ptok:
            continue
        if all(t in TOPIC_GENERIC_TOKENS for t in ptok):
            continue
        if any(jaccard(ptok, prev) >= 0.7 for prev in phrase_tokens_acc):
            continue
        seen.add(folded)
        phrases.append(p)
        phrase_tokens_acc.append(ptok)
        if len(phrases) >= 4:
            break
    if phrases:
        return phrases
    return extract_keyphrases_heuristic(source_text)


def extract_keyphrases_heuristic(source_text):
    source = clean_text(source_text, 1800)
    tokens = [t for t in word_tokens(source, lower=True) if t not in TOPIC_STOPWORDS and len(t) >= 4]
    if not tokens:
        return []

    unigram_counts = Counter(tokens)
    bigrams = [f"{tokens[i]} {tokens[i+1]}" for i in range(len(tokens) - 1)]
    bigram_counts = Counter(bigrams)

    scored = []
    for phrase, count in bigram_counts.items():
        ptok = phrase_tokens(phrase)
        if not ptok or all(t in TOPIC_GENERIC_TOKENS for t in ptok):
            continue
        score = count * 2 + len(ptok)
        scored.append((phrase, score))
    for phrase, count in unigram_counts.items():
        ptok = phrase_tokens(phrase)
        if not ptok or all(t in TOPIC_GENERIC_TOKENS for t in ptok):
            continue
        score = count
        scored.append((phrase, score))

    scored.sort(key=lambda item: item[1], reverse=True)
    out = []
    out_tokens = []
    for phrase, _score in scored:
        ptok = phrase_tokens(phrase)
        if any(jaccard(ptok, prev) >= 0.75 for prev in out_tokens):
            continue
        out.append(phrase)
        out_tokens.append(ptok)
        if len(out) >= 4:
            break
    return out


def build_generated_context_from_nlp(label, lead, keyphrases):
    action_map = {
        "Wspolpraca i umowy": "Post informuje o nowej wspolpracy lub uzgodnieniu miedzy firmami.",
        "Webinar lub szkolenie": "Post zaprasza na szkolenie lub webinar i wskazuje jego zakres.",
        "Wydarzenie branzowe": "Post dotyczy obecnosci firmy na wydarzeniu branzowym.",
        "Promocja lub konkurs": "Post opisuje akcje promocyjna lub konkurs i warunki udzialu.",
        "Nowy produkt lub rozwiazanie": "Post prezentuje nowe rozwiazanie oraz jego zastosowanie.",
        "Nagroda lub wyroznienie": "Post komunikuje otrzymane wyroznienie i jego kontekst.",
        "Rekrutacja i kariera": "Post dotyczy naboru i warunkow pracy na wskazane stanowiska.",
        "Poradnik i wskazowki": "Post zawiera praktyczne wskazowki dotyczace konkretnego problemu.",
        "Aktualizacja firmowa": "Post przekazuje biezaca aktualizacje dzialan firmy.",
    }
    if lead:
        sentence = clean_generated_sentence(lead, max_words=24)
        if keyphrases:
            details = ", ".join(keyphrases[:2])
            return clean_generated_sentence(f"{sentence}. Najwazniejsze elementy: {details}.", max_words=30)
        return sentence
    details = ", ".join(keyphrases[:3]) if keyphrases else TOPIC_DEFAULTS.get(label, "aktualizacja firmy")
    return clean_generated_sentence(f"{action_map.get(label)} Dotyczy: {details}.", max_words=30)


def build_ai_topic_context(raw_text, label):
    source = clean_text(raw_text, 1800)
    source = re.sub(r"https?://\S+", " ", source)
    source = re.sub(r"[@#][\w\-]+", " ", source)
    source = re.sub(r"\s+", " ", source).strip(" -,:;")
    if len(source) < 30:
        return ""
    if looks_like_template_source(source):
        return ""

    model, kw_model = init_topic_ai()
    lead = pick_lead_sentence(model, source)
    keyphrases = extract_keyphrases(kw_model, source)
    lead_tokens = word_tokens(lead)
    short_ratio = (sum(1 for t in lead_tokens if len(t) == 1) / max(1, len(lead_tokens))) if lead_tokens else 1.0
    if len(lead_tokens) < 4 or short_ratio > 0.4:
        lead = ""
    if not lead and not keyphrases:
        return ""
    return build_generated_context_from_nlp(label, lead, keyphrases)


def build_fallback_topic_context(raw, label):
    brief = clean_text(raw, 500)
    brief = re.sub(r"https?://\S+", " ", brief)
    brief = re.sub(r"[@#][\w\-]+", " ", brief)
    brief = re.sub(r"\s+", " ", brief).strip(" -,:;")

    sentence = split_sentences(brief)
    if sentence:
        lead = clean_generated_sentence(sentence[0], max_words=22)
        if len(word_tokens(lead)) >= 8:
            return lead

    brief_words = word_tokens(brief)
    if len(brief_words) >= 8:
        return clean_generated_sentence(" ".join(brief_words[:18]), max_words=18)
    return TOPIC_DEFAULTS.get(label, "najwazniejsza informacja z publikacji")


def parse_int(value):
    if value is None:
        return None
    digits = re.sub(r"[^\d]", "", str(value))
    if not digits:
        return None
    try:
        return int(digits)
    except ValueError:
        return None


def parse_iso_dt(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except Exception:
        return None


def dt_to_date_time(dt):
    if dt is None:
        return None, None
    dt = dt.astimezone(timezone.utc)
    return dt.date(), dt.time().replace(microsecond=0)


def within_days(dt, days, now_utc):
    if dt is None or days is None:
        return True
    return dt >= (now_utc - timedelta(days=days))


def find_or_add_header(ws, header_name):
    header_l = header_name.strip().lower()
    for col in range(1, max(1, ws.max_column) + 1):
        if normalize_header(ws.cell(row=1, column=col).value) == header_l:
            return col
    for col in range(1, max(1, ws.max_column) + 2):
        val = ws.cell(row=1, column=col).value
        if val is None or str(val).strip() == "":
            ws.cell(row=1, column=col, value=header_name)
            return col
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value=header_name)
    return col


def header_map(ws):
    out = {}
    for col in range(1, max(1, ws.max_column) + 1):
        h = normalize_header(ws.cell(row=1, column=col).value)
        if h:
            out[h] = col
    return out


def ensure_topic_headers(ws):
    for name in ["topic", "raw_text", "topic_quality", "topic_reason", "scrape_date"]:
        find_or_add_header(ws, name)
    return header_map(ws)


def ensure_scrape_headers_in_workbook(wb, sheet_names=None):
    target_names = sheet_names or ["Facebook", "Linkedin", "YouTube", "TikTok", "Instagram"]
    for name in target_names:
        if name in wb.sheetnames:
            ensure_topic_headers(wb[name])


def find_insert_row(ws, key_col=None):
    start = 2
    end = max(2, ws.max_row)
    for row in range(start, end + 1):
        key_val = ws.cell(row=row, column=key_col).value if key_col else None
        if key_col and key_val not in (None, ""):
            continue
        is_empty = True
        for col in range(1, max(1, ws.max_column) + 1):
            if ws.cell(row=row, column=col).value not in (None, ""):
                is_empty = False
                break
        if is_empty:
            return row
    return ws.max_row + 1


def append_row(ws, hm, payload, key_name=None):
    key_col = hm.get(normalize_header(key_name)) if key_name else None
    row = find_insert_row(ws, key_col)
    for key, value in payload.items():
        col = hm.get(normalize_header(key))
        if col:
            ws.cell(row=row, column=col, value=value)


class HttpClient:
    def __init__(self):
        self.ctx = ssl.create_default_context()
        self.default_headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "en-US,en;q=0.9",
        }

    def open(self, url, headers=None, opener=None, timeout=35):
        hdr = dict(self.default_headers)
        if headers:
            hdr.update(headers)
        req = urllib.request.Request(url, headers=hdr)
        if opener is not None:
            with opener.open(req, timeout=timeout) as resp:
                return resp.status, resp.read(), dict(resp.headers)
        with urllib.request.urlopen(req, timeout=timeout, context=self.ctx) as resp:
            return resp.status, resp.read(), dict(resp.headers)

    def text(self, url, headers=None, opener=None, timeout=35):
        status, body, resp_headers = self.open(url, headers=headers, opener=opener, timeout=timeout)
        return status, body.decode("utf-8", "ignore"), resp_headers

    def json(self, url, headers=None, opener=None, timeout=35):
        status, body, resp_headers = self.open(url, headers=headers, opener=opener, timeout=timeout)
        return status, json.loads(body), resp_headers


def get_source_profiles(wb, source_sheet_name):
    if source_sheet_name not in wb.sheetnames:
        return []
    ws = wb[source_sheet_name]
    hm = header_map(ws)
    c_col = hm.get("company")
    u_col = hm.get("profile_url")
    if not c_col or not u_col:
        return []
    out = []
    for row in range(2, ws.max_row + 1):
        company = clean_text(ws.cell(row=row, column=c_col).value, 200)
        url = clean_text(ws.cell(row=row, column=u_col).value, 500)
        if company and url:
            out.append((company, url))
    return out


def existing_keys(ws, key_col):
    keys = set()
    if key_col is None:
        return keys
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=key_col).value
        if v:
            keys.add(clean_text(v, 500))
    return keys


def remove_rows_matching(ws, predicate):
    rows = []
    for row in range(2, ws.max_row + 1):
        if predicate(row):
            rows.append(row)
    for row in reversed(rows):
        ws.delete_rows(row, 1)
    return len(rows)


def compact_sheet(ws):
    records = []
    for row in range(2, ws.max_row + 1):
        vals = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        if any(v not in (None, "") for v in vals):
            records.append(vals)
    # Clear data rows
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col, value=None)
    # Write compacted data
    write_row = 2
    for vals in records:
        for col, val in enumerate(vals, start=1):
            ws.cell(row=write_row, column=col, value=val)
        write_row += 1


def cleanup_workbook_layout(wb):
    removed = {}
    if "TikTok" in wb.sheetnames:
        ws = wb["TikTok"]
        hm = header_map(ws)
        post_col = hm.get("post_url")
        if post_col:
            removed["TikTok_snapshots"] = remove_rows_matching(
                ws,
                lambda r: "snapshot=" in clean_text(ws.cell(row=r, column=post_col).value, 500),
            )
    if "Facebook" in wb.sheetnames:
        ws = wb["Facebook"]
        hm = header_map(ws)
        post_col = hm.get("post_url")
        if post_col:
            removed["Facebook_snapshots"] = remove_rows_matching(
                ws,
                lambda r: "snapshot=" in clean_text(ws.cell(row=r, column=post_col).value, 500),
            )
    if "Linkedin" in wb.sheetnames:
        compact_sheet(wb["Linkedin"])
    return removed


def extract_post_text_from_html(html_text):
    if not html_text:
        return ""
    patterns = [
        r'<meta\s+property="og:description"\s+content="([^"]+)"',
        r'<meta\s+name="description"\s+content="([^"]+)"',
        r'<meta\s+name="twitter:description"\s+content="([^"]+)"',
    ]
    for pat in patterns:
        m = re.search(pat, html_text, re.I)
        if m:
            candidate = clean_text(m.group(1), 1800)
            candidate = strip_tags(candidate)
            if len(word_tokens(candidate)) >= 6:
                return candidate

    script_desc = re.search(r'"description"\s*:\s*"\s*([^"]+?)\s*"', html_text, re.I)
    if script_desc:
        candidate = clean_text(script_desc.group(1), 1800)
        candidate = strip_tags(candidate)
        if len(word_tokens(candidate)) >= 6:
            return candidate

    title_m = re.search(r"<title[^>]*>(.*?)</title>", html_text, re.I | re.S)
    if title_m:
        title = strip_tags(title_m.group(1))
        if len(word_tokens(title)) >= 6:
            return clean_text(title, 300)
    return ""


def fetch_raw_text_from_url(post_url, http_client):
    url = clean_text(post_url, 800)
    if not url or http_client is None:
        return ""
    if url in _RAW_TEXT_URL_CACHE:
        return _RAW_TEXT_URL_CACHE[url]

    lower_url = url.lower()
    if "tiktok.com/" in lower_url and YoutubeDL is not None:
        try:
            details = ydl_extract(url, flat=False)
            candidate = clean_text((details or {}).get("description") or (details or {}).get("title"), 1800)
            if len(word_tokens(candidate)) >= 5:
                _RAW_TEXT_URL_CACHE[url] = candidate
                return candidate
        except Exception:
            pass

    try:
        _status, html_text, _ = http_client.text(
            url,
            headers={"Accept-Language": "en-US,en;q=0.9", "User-Agent": "Mozilla/5.0"},
            timeout=15,
        )
        extracted = extract_post_text_from_html(html_text)
    except Exception:
        extracted = ""
    _RAW_TEXT_URL_CACHE[url] = extracted
    return extracted


def rewrite_topic_column(ws, http_client=None, backfill_from_urls=False):
    hm = ensure_topic_headers(ws)
    topic_col = hm.get("topic")
    raw_col = hm.get("raw_text")
    quality_col = hm.get("topic_quality")
    reason_col = hm.get("topic_reason")
    url_col = hm.get("post_url") or hm.get("vid_url")
    changed = 0
    for row in range(2, ws.max_row + 1):
        has_data = False
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col).value not in (None, ""):
                has_data = True
                break
        if not has_data:
            continue

        raw_value = normalize_raw_text_content(ws.cell(row=row, column=raw_col).value) if raw_col else ""
        post_url = ws.cell(row=row, column=url_col).value if url_col else ""
        current_topic = clean_text(ws.cell(row=row, column=topic_col).value, 300) if topic_col else ""
        current_quality = clean_text(ws.cell(row=row, column=quality_col).value, 30).lower() if quality_col else ""
        current_reason = clean_text(ws.cell(row=row, column=reason_col).value, 300) if reason_col else ""

        if not raw_value and backfill_from_urls and post_url:
            fetched = fetch_raw_text_from_url(post_url, http_client)
            if fetched:
                raw_value = normalize_raw_text_content(fetched)
                if raw_col:
                    ws.cell(row=row, column=raw_col, value=raw_value)
                    changed += 1

        if raw_value:
            new_topic, new_quality, new_reason = summarize_topic_with_meta(raw_value, post_url=post_url)
        else:
            new_topic = current_topic
            new_quality = current_quality or "low"
            new_reason = current_reason or "Brak raw_text - temat nie zostal przepisany przez AI."

        if topic_col and new_topic and current_topic != clean_text(new_topic, 300):
            ws.cell(row=row, column=topic_col, value=new_topic)
            changed += 1
        if quality_col and current_quality != clean_text(new_quality, 30).lower():
            ws.cell(row=row, column=quality_col, value=new_quality)
            changed += 1
        if reason_col and current_reason != clean_text(new_reason, 300):
            ws.cell(row=row, column=reason_col, value=new_reason)
            changed += 1
    return changed


def normalize_topics_in_workbook(wb, http_client=None, backfill_from_urls=False, sheet_names=None):
    out = {}
    target_names = sheet_names or ["Facebook", "Linkedin", "YouTube", "TikTok", "Instagram"]
    for name in target_names:
        if name in wb.sheetnames:
            out[name] = rewrite_topic_column(wb[name], http_client=http_client, backfill_from_urls=backfill_from_urls)
    return out


def parse_li_relative_time(text, now_utc):
    t = clean_text(text).lower().replace(" ", "")
    m = re.match(r"(\d+)([a-z]+)", t)
    if not m:
        return None
    n = int(m.group(1))
    u = m.group(2)
    if u in ("m", "min"):
        return now_utc - timedelta(minutes=n)
    if u in ("h", "hr", "hrs"):
        return now_utc - timedelta(hours=n)
    if u in ("d", "day", "days"):
        return now_utc - timedelta(days=n)
    if u in ("w", "wk", "wks"):
        return now_utc - timedelta(weeks=n)
    if u in ("mo", "mos", "month", "months"):
        return now_utc - timedelta(days=n * 30)
    if u in ("y", "yr", "yrs", "year", "years"):
        return now_utc - timedelta(days=n * 365)
    return None


def parse_ig_username(profile_url):
    m = re.search(r"instagram\.com/([^/?#]+)/?", profile_url, re.IGNORECASE)
    return m.group(1) if m else None


def parse_li_slug(profile_url):
    m = re.search(r"linkedin\.com/company/([^/?#]+)", profile_url, re.IGNORECASE)
    return m.group(1) if m else None


def parse_tt_username(profile_url):
    m = re.search(r"tiktok\.com/@([^/?#]+)", profile_url, re.IGNORECASE)
    return m.group(1) if m else None


def parse_fb_slug(profile_url):
    m = re.search(r"facebook\.com/([^/?#]+)", profile_url, re.IGNORECASE)
    return m.group(1) if m else None


def meta_config_path():
    env_path = clean_text(os.getenv("META_CONFIG_PATH", ""), 500)
    if env_path:
        return Path(env_path)
    return topic_project_root() / "meta_api_config.json"


def get_meta_config():
    global _META_CONFIG_CACHE
    if _META_CONFIG_CACHE is not None:
        return _META_CONFIG_CACHE

    cfg = {}
    cfg_path = meta_config_path()
    if cfg_path.exists():
        try:
            loaded = json.loads(cfg_path.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                cfg.update(loaded)
        except Exception:
            pass

    env_overrides = {
        "access_token": clean_text(os.getenv("META_ACCESS_TOKEN", ""), 8000),
        "graph_api_version": clean_text(os.getenv("META_GRAPH_API_VERSION", ""), 20),
        "instagram_business_account_id": clean_text(os.getenv("META_IG_BUSINESS_ACCOUNT_ID", ""), 100),
        "facebook_enabled": clean_text(os.getenv("META_FACEBOOK_ENABLED", ""), 20),
        "instagram_enabled": clean_text(os.getenv("META_INSTAGRAM_ENABLED", ""), 20),
    }
    for key, value in env_overrides.items():
        if value:
            cfg[key] = value

    _META_CONFIG_CACHE = cfg
    return _META_CONFIG_CACHE


def meta_config_flag(name, default=False):
    value = get_meta_config().get(name)
    if isinstance(value, bool):
        return value
    value = clean_text(value, 20).lower()
    if not value:
        return default
    return value in ("1", "true", "yes", "y", "on")


def meta_access_token():
    return clean_text(get_meta_config().get("access_token"), 8000)


def meta_graph_api_version():
    version = clean_text(get_meta_config().get("graph_api_version"), 20)
    return version or "v21.0"


def meta_instagram_business_account_id():
    return clean_text(get_meta_config().get("instagram_business_account_id"), 100)


def meta_facebook_enabled():
    if not meta_access_token():
        return False
    return meta_config_flag("facebook_enabled", default=True)


def meta_instagram_enabled():
    if not meta_access_token() or not meta_instagram_business_account_id():
        return False
    return meta_config_flag("instagram_enabled", default=True)


def build_meta_graph_url(path, params=None):
    path_part = clean_text(path, 500).lstrip("/")
    query = dict(params or {})
    query["access_token"] = meta_access_token()
    return f"https://graph.facebook.com/{meta_graph_api_version()}/{path_part}?{urllib.parse.urlencode(query, doseq=True)}"


def meta_graph_json(http_client, path, params=None, timeout=45):
    url = build_meta_graph_url(path, params=params)
    _status, data, _headers = http_client.json(url, timeout=timeout)
    return data


def graph_summary_total(node):
    if not isinstance(node, dict):
        return None
    summary = node.get("summary") or {}
    return parse_int(summary.get("total_count"))


def join_meta_text_parts(*parts):
    out = []
    for part in parts:
        txt = normalize_raw_text_content(part)
        if txt and txt not in out:
            out.append(txt)
    return clean_text(" | ".join(out), 1800)


def fb_post_text_from_graph(post):
    attachments = ((post or {}).get("attachments") or {}).get("data") or []
    attachment_parts = []
    for item in attachments[:3]:
        attachment_parts.append(item.get("title"))
        attachment_parts.append(item.get("description"))
        attachment_parts.append(item.get("url"))
    return join_meta_text_parts(
        (post or {}).get("message"),
        (post or {}).get("story"),
        *attachment_parts,
    )


def scrape_instagram_via_meta_api(wb, http_client, now_utc, days, max_per_profile, scrape_date):
    if not meta_instagram_enabled():
        return None

    ws = wb["Instagram"]
    hm = ensure_topic_headers(ws)
    key_col = hm.get("post_url")
    keys = existing_keys(ws, key_col)

    profiles = get_source_profiles(wb, "Source_ig")
    added = 0
    ig_business_id = meta_instagram_business_account_id()

    for company, profile_url in profiles:
        username = parse_ig_username(profile_url)
        if not username:
            continue
        try:
            fields = (
                f'business_discovery.username({username})'
                "{media.limit("
                f"{max_per_profile}"
                "){caption,comments_count,id,like_count,media_type,permalink,timestamp}}"
            )
            data = meta_graph_json(http_client, ig_business_id, params={"fields": fields})
            discovery = (data or {}).get("business_discovery") or {}
            media_items = discovery.get("media") or {}
            for item in (media_items.get("data") or [])[:max_per_profile]:
                post_url = clean_text(item.get("permalink"), 500)
                if not post_url or post_url in keys:
                    continue
                post_dt = parse_iso_dt(item.get("timestamp"))
                if not within_days(post_dt, days, now_utc):
                    continue
                post_date, post_time = dt_to_date_time(post_dt)
                caption = clean_text(item.get("caption"), 1800)
                topic_fields = build_topic_payload(caption, post_url=post_url)
                payload = {
                    "post_url": post_url,
                    "company": company,
                    "likes": parse_int(item.get("like_count")),
                    "comments": parse_int(item.get("comments_count")),
                    "shares": None,
                    "reach": None,
                    "post_date": post_date,
                    "post_time": post_time,
                    "scrape_date": scrape_date,
                }
                payload.update(topic_fields)
                append_row(ws, hm, payload, key_name="post_url")
                keys.add(post_url)
                added += 1
        except Exception:
            continue
    return {"profiles": len(profiles), "added": added, "source": "meta_api"}


def scrape_facebook_via_meta_api(wb, http_client, now_utc, days, max_per_profile, scrape_date):
    if not meta_facebook_enabled():
        return None

    ws = wb["Facebook"]
    hm = ensure_topic_headers(ws)
    key_col = hm.get("post_url")
    keys = existing_keys(ws, key_col)

    profiles = get_source_profiles(wb, "Source_fb")
    added = 0
    for company, profile_url in profiles:
        page_slug = parse_fb_slug(profile_url)
        if not page_slug:
            continue
        try:
            fields = (
                "attachments{description,title,url},"
                "comments.summary(true).limit(0),"
                "created_time,"
                "message,"
                "permalink_url,"
                "reactions.summary(true).limit(0),"
                "shares,"
                "story"
            )
            data = meta_graph_json(
                http_client,
                f"{page_slug}/posts",
                params={"fields": fields, "limit": max_per_profile},
            )
            for post in (data.get("data") or [])[:max_per_profile]:
                post_url = clean_text(post.get("permalink_url"), 500)
                if not post_url:
                    post_id = clean_text(post.get("id"), 200)
                    if post_id:
                        post_url = f"https://www.facebook.com/{post_id}"
                if not post_url or post_url in keys:
                    continue
                post_dt = parse_iso_dt(post.get("created_time"))
                if not within_days(post_dt, days, now_utc):
                    continue
                post_date, post_time = dt_to_date_time(post_dt)
                raw_text = fb_post_text_from_graph(post)
                topic_fields = build_topic_payload(raw_text, post_url=post_url)
                payload = {
                    "post_url": post_url,
                    "company": company,
                    "likes": graph_summary_total(post.get("reactions")),
                    "comments": graph_summary_total(post.get("comments")),
                    "shares": parse_int((post.get("shares") or {}).get("count")),
                    "reach": None,
                    "post_date": post_date,
                    "post_time": post_time,
                    "scrape_date": scrape_date,
                }
                payload.update(topic_fields)
                append_row(ws, hm, payload, key_name="post_url")
                keys.add(post_url)
                added += 1
        except Exception:
            continue
    return {"profiles": len(profiles), "added": added, "source": "meta_api"}


def parse_yt_channel_id(html_text):
    m = re.search(r'"externalId":"(UC[^"]+)"', html_text)
    if m:
        return m.group(1)
    m = re.search(r'"browseId":"(UC[^"]+)"', html_text)
    if m:
        return m.group(1)
    return None


def fetch_yt_metrics(http_client, video_url):
    try:
        _status, html_text, _ = http_client.text(video_url)
        view_count = None
        likes = None
        comments = None

        m = re.search(r'"viewCount":"(\d+)"', html_text)
        if m:
            view_count = int(m.group(1))
        m = re.search(r'"likeCount":"(\d+)"', html_text)
        if m:
            likes = int(m.group(1))
        m = re.search(r'"commentCount":"(\d+)"', html_text)
        if m:
            comments = int(m.group(1))
        return {"views": view_count, "likes": likes, "comments": comments}
    except Exception:
        return {"views": None, "likes": None, "comments": None}


def extract_universal_json(html_text):
    m = re.search(
        r'<script id="__UNIVERSAL_DATA_FOR_REHYDRATION__" type="application/json">(.*?)</script>',
        html_text,
        re.S,
    )
    if not m:
        return None
    try:
        return json.loads(m.group(1))
    except Exception:
        return None


def scrape_youtube(wb, http_client, now_utc, days, max_per_profile, scrape_date):
    ws = wb["YouTube"]
    find_or_add_header(ws, "views")
    hm = ensure_topic_headers(ws)
    key_col = hm.get("vid_url") or hm.get("post_url")
    keys = existing_keys(ws, key_col)

    profiles = get_source_profiles(wb, "Source_yt")
    added = 0
    for company, profile_url in profiles:
        try:
            status, videos_html, _ = http_client.text(profile_url.rstrip("/") + "/videos")
            if status != 200:
                continue
            channel_id = parse_yt_channel_id(videos_html)
            if not channel_id:
                continue
            feed_url = f"https://www.youtube.com/feeds/videos.xml?channel_id={channel_id}"
            _status_f, feed_bytes, _ = http_client.open(feed_url)
            root = ET.fromstring(feed_bytes)
            entries = root.findall("atom:entry", YT_NS)[:max_per_profile]
            for e in entries:
                vid = clean_text(e.findtext("yt:videoId", default="", namespaces=YT_NS))
                if not vid:
                    continue
                post_url = f"https://www.youtube.com/watch?v={vid}"
                if post_url in keys:
                    continue
                published = parse_iso_dt(e.findtext("atom:published", default="", namespaces=YT_NS))
                if not within_days(published, days, now_utc):
                    continue
                post_date, post_time = dt_to_date_time(published)
                title = clean_text(e.findtext("atom:title", default="", namespaces=YT_NS), 500)
                metrics = fetch_yt_metrics(http_client, post_url)
                topic_fields = build_topic_payload(title, post_url=post_url)
                payload = {
                    "vid_url": post_url,
                    "post_url": post_url,
                    "company": company,
                    "likes": metrics.get("likes"),
                    "comments": metrics.get("comments"),
                    "shares": None,
                    "reach": None,
                    "post_date": post_date,
                    "post_time": post_time,
                    "views": metrics.get("views"),
                    "scrape_date": scrape_date,
                }
                payload.update(topic_fields)
                append_row(ws, hm, payload, key_name="vid_url")
                keys.add(post_url)
                added += 1
        except Exception:
            continue
    return {"profiles": len(profiles), "added": added}


def scrape_instagram_legacy(wb, http_client, now_utc, days, max_per_profile, scrape_date):
    ws = wb["Instagram"]
    hm = ensure_topic_headers(ws)
    key_col = hm.get("post_url")
    keys = existing_keys(ws, key_col)

    profiles = get_source_profiles(wb, "Source_ig")
    added = 0
    for company, profile_url in profiles:
        username = parse_ig_username(profile_url)
        if not username:
            continue
        try:
            jar = http.cookiejar.CookieJar()
            opener = urllib.request.build_opener(
                urllib.request.HTTPCookieProcessor(jar),
                urllib.request.HTTPSHandler(context=http_client.ctx),
            )
            profile_page = f"https://www.instagram.com/{username}/"
            _status, html_text, _ = http_client.text(
                profile_page,
                opener=opener,
                headers={"User-Agent": "Mozilla/5.0", "Accept-Language": "en-US,en;q=0.9"},
            )
            uid_match = re.search(r'"user_id":"(\d+)"', html_text)
            if not uid_match:
                continue
            uid = uid_match.group(1)
            csrf = ""
            for c in jar:
                if c.name == "csrftoken":
                    csrf = c.value
                    break
            headers = {
                "X-IG-App-ID": "936619743392459",
                "X-CSRFToken": csrf,
                "Referer": profile_page,
                "Accept": "*/*",
                "User-Agent": "Mozilla/5.0",
            }
            feed_url = f"https://www.instagram.com/api/v1/feed/user/{uid}/?count={max_per_profile}"
            _status_j, feed_json, _ = http_client.json(feed_url, headers=headers, opener=opener)
            for item in feed_json.get("items", [])[:max_per_profile]:
                code = clean_text(item.get("code"), 100)
                if not code:
                    continue
                post_url = f"https://www.instagram.com/p/{code}/"
                if post_url in keys:
                    continue
                ts = item.get("taken_at")
                post_dt = datetime.fromtimestamp(ts, tz=timezone.utc) if ts else None
                if not within_days(post_dt, days, now_utc):
                    continue
                post_date, post_time = dt_to_date_time(post_dt)
                caption_obj = item.get("caption") or {}
                caption = clean_text(caption_obj.get("text", ""), 500)
                topic_fields = build_topic_payload(caption, post_url=post_url)
                payload = {
                    "post_url": post_url,
                    "company": company,
                    "likes": item.get("like_count"),
                    "comments": item.get("comment_count"),
                    "shares": None,
                    "reach": item.get("play_count") or item.get("view_count"),
                    "post_date": post_date,
                    "post_time": post_time,
                    "scrape_date": scrape_date,
                }
                payload.update(topic_fields)
                append_row(ws, hm, payload, key_name="post_url")
                keys.add(post_url)
                added += 1
        except Exception:
            continue
    return {"profiles": len(profiles), "added": added, "source": "legacy"}


def scrape_linkedin(wb, http_client, now_utc, days, max_per_profile, scrape_date):
    ws = wb["Linkedin"]
    hm = ensure_topic_headers(ws)
    key_col = hm.get("post_id") or hm.get("post_url")
    keys = existing_keys(ws, key_col)

    profiles = get_source_profiles(wb, "Source_li")
    added = 0
    for company, profile_url in profiles:
        slug = parse_li_slug(profile_url)
        if not slug:
            continue
        try:
            page_url = f"https://www.linkedin.com/company/{slug}/"
            _status, html_text, _ = http_client.text(page_url)
            blocks = re.findall(r'(<div data-id="entire-feed-card-link"[\s\S]*?</article>)', html_text)[:max_per_profile]
            for block in blocks:
                aid_m = re.search(r'data-activity-urn="urn:li:activity:(\d+)"', block)
                if not aid_m:
                    continue
                activity_id = aid_m.group(1)
                key = activity_id
                if key in keys:
                    continue

                url_m = re.search(r'href="([^"]*activity-' + re.escape(activity_id) + r'[^"]*)"', block)
                if not url_m:
                    url_m = re.search(r'href="([^"]*activity-[^"]*)"', block)
                post_url = html.unescape(url_m.group(1)) if url_m else ""
                if post_url.startswith("/"):
                    post_url = "https://www.linkedin.com" + post_url

                likes = parse_int(re.search(r'data-num-reactions="(\d+)"', block).group(1)) if re.search(
                    r'data-num-reactions="(\d+)"', block
                ) else None
                comments = parse_int(re.search(r'data-num-comments="(\d+)"', block).group(1)) if re.search(
                    r'data-num-comments="(\d+)"', block
                ) else None
                rel_time_raw = strip_tags(re.search(r"<time[^>]*>([\s\S]*?)</time>", block).group(1)) if re.search(
                    r"<time[^>]*>([\s\S]*?)</time>", block
                ) else ""
                dt_guess = parse_li_relative_time(rel_time_raw, now_utc)
                if not within_days(dt_guess, days, now_utc):
                    continue
                post_date, post_time = dt_to_date_time(dt_guess)

                text_m = re.search(
                    r'<p class="attributed-text-segment-list__content[^"]*"[^>]*>([\s\S]*?)</p>',
                    block,
                )
                post_text = strip_tags(text_m.group(1)) if text_m else ""
                topic_fields = build_topic_payload(post_text, post_url=post_url)

                payload = {
                    "post_url": post_url,
                    "post_id": activity_id,
                    "company": company,
                    "likes": likes,
                    "comments": comments,
                    "shares": None,
                    "reach": None,
                    "post_date": post_date,
                    "post_time": post_time,
                    "scrape_date": scrape_date,
                }
                payload.update(topic_fields)
                append_row(ws, hm, payload, key_name="post_id")
                keys.add(key)
                added += 1
        except Exception:
            continue
    return {"profiles": len(profiles), "added": added}


def ydl_extract(url, flat=False):
    if YoutubeDL is None:
        return None
    opts = {"quiet": True, "skip_download": True, "no_warnings": True}
    if flat:
        opts["extract_flat"] = "in_playlist"
    with YoutubeDL(opts) as ydl:
        return ydl.extract_info(url, download=False)


def scrape_tiktok(wb, http_client, now_utc, days, max_per_profile, scrape_date):
    ws = wb["TikTok"]
    hm = ensure_topic_headers(ws)
    key_col = hm.get("post_url")
    keys = existing_keys(ws, key_col)

    profiles = get_source_profiles(wb, "Source_tt")
    added = 0
    if YoutubeDL is None:
        return {"profiles": len(profiles), "added": 0}

    for company, profile_url in profiles:
        username = parse_tt_username(profile_url)
        if not username:
            continue
        try:
            playlist = ydl_extract(f"https://www.tiktok.com/@{username}", flat=True)
            entries = (playlist or {}).get("entries") or []

            added_for_profile = 0
            for entry in entries:
                if added_for_profile >= max_per_profile:
                    break
                video_url = clean_text(entry.get("url") or entry.get("webpage_url"), 500)
                if not video_url or "/video/" not in video_url:
                    continue
                if video_url in keys:
                    continue
                details = ydl_extract(video_url, flat=False)
                if not details:
                    continue
                ts = details.get("timestamp")
                dt = datetime.fromtimestamp(ts, tz=timezone.utc) if ts else None
                if not within_days(dt, days, now_utc):
                    continue
                post_date, post_time = dt_to_date_time(dt)
                raw_caption = details.get("description") or details.get("title")
                topic_fields = build_topic_payload(raw_caption, post_url=video_url)
                payload = {
                    "post_url": video_url,
                    "company": company,
                    "likes": details.get("like_count"),
                    "comments": details.get("comment_count"),
                    "shares": details.get("repost_count"),
                    "reach": details.get("view_count"),
                    "post_date": post_date,
                    "post_time": post_time,
                    "scrape_date": scrape_date,
                }
                payload.update(topic_fields)
                append_row(ws, hm, payload, key_name="post_url")
                keys.add(video_url)
                added += 1
                added_for_profile += 1
        except Exception:
            continue
    return {"profiles": len(profiles), "added": added}


def scrape_facebook_legacy(wb, http_client, now_utc, scrape_date, days=7, max_per_profile=5):
    ws = wb["Facebook"]
    hm = ensure_topic_headers(ws)
    key_col = hm.get("post_url")
    keys = existing_keys(ws, key_col)

    profiles = get_source_profiles(wb, "Source_fb")
    added = 0
    for company, profile_url in profiles:
        slug = parse_fb_slug(profile_url)
        if not slug:
            continue
        try:
            page_url = f"https://www.facebook.com/{slug}/"
            _status, html_text, _ = http_client.text(
                page_url,
                headers={"Accept-Language": "en-US,en;q=0.9", "User-Agent": "Mozilla/5.0"},
            )
            post_links = set()
            for pat in [
                r'https://www\.facebook\.com/[^"\s<>]*(?:/posts/[^"\s<>]+|/reel/\d+|/videos/\d+|story\.php\?story_fbid=[^"\s<>]+)',
                r'https:\\/\\/www\.facebook\.com\\/[^"\\]*(?:\\/posts\\/[^"\\]+|\\/reel\\/\\d+|\\/videos\\/\\d+|story\.php\\?story_fbid=[^"\\]+)',
            ]:
                for m in re.finditer(pat, html_text):
                    u = m.group(0).replace("\\/", "/")
                    u = clean_text(u, 500)
                    if "/posts/" in u or "/reel/" in u or "/videos/" in u or "story.php?story_fbid=" in u:
                        post_links.add(u)

            for post_url in list(post_links)[:max_per_profile]:
                if post_url in keys:
                    continue
                try:
                    _sp, post_html, _ = http_client.text(
                        post_url,
                        headers={"Accept-Language": "en-US,en;q=0.9", "User-Agent": "Mozilla/5.0"},
                    )
                except Exception:
                    post_html = ""

                raw_post_text = ""
                post_dt = None
                if post_html:
                    mt = re.search(r'<meta property="og:description" content="([^"]+)"', post_html, re.I)
                    raw_post_text = clean_text(mt.group(1), 500) if mt else ""
                    md = re.search(r'<meta property="article:published_time" content="([^"]+)"', post_html, re.I)
                    if md:
                        post_dt = parse_iso_dt(md.group(1))
                if not within_days(post_dt, days, now_utc):
                    continue
                post_date, post_time = dt_to_date_time(post_dt)
                topic_fields = build_topic_payload(raw_post_text, post_url=post_url)

                payload = {
                    "post_url": post_url,
                    "company": company,
                    "likes": None,
                    "comments": None,
                    "shares": None,
                    "reach": None,
                    "post_date": post_date,
                    "post_time": post_time,
                    "scrape_date": scrape_date,
                }
                payload.update(topic_fields)
                append_row(ws, hm, payload, key_name="post_url")
                keys.add(post_url)
                added += 1
        except Exception:
            continue
    return {"profiles": len(profiles), "added": added, "source": "legacy"}


def scrape_instagram(wb, http_client, now_utc, days, max_per_profile, scrape_date):
    api_result = scrape_instagram_via_meta_api(wb, http_client, now_utc, days, max_per_profile, scrape_date)
    if api_result is not None:
        return api_result
    return scrape_instagram_legacy(wb, http_client, now_utc, days, max_per_profile, scrape_date)


def scrape_facebook(wb, http_client, now_utc, scrape_date, days=7, max_per_profile=5):
    api_result = scrape_facebook_via_meta_api(wb, http_client, now_utc, days, max_per_profile, scrape_date)
    if api_result is not None:
        return api_result
    return scrape_facebook_legacy(wb, http_client, now_utc, scrape_date, days=days, max_per_profile=max_per_profile)


def normalize_platform_key(value):
    token = re.sub(r"[^a-z0-9]+", "", ascii_fold(value))
    aliases = {
        "facebook": "facebook",
        "fb": "facebook",
        "linkedin": "linkedin",
        "linkedin": "linkedin",
        "li": "linkedin",
        "youtube": "youtube",
        "yt": "youtube",
        "instagram": "instagram",
        "ig": "instagram",
        "tiktok": "tiktok",
        "tt": "tiktok",
    }
    return aliases.get(token, "")


def get_platform_specs():
    return [
        {
            "key": "youtube",
            "label": "YouTube",
            "sheet": "YouTube",
            "source_sheet": "Source_yt",
            "run": lambda wb, http_client, now_utc, args: scrape_youtube(
                wb, http_client, now_utc, args.days, args.max_per_profile, args.scrape_date
            ),
        },
        {
            "key": "instagram",
            "label": "Instagram",
            "sheet": "Instagram",
            "source_sheet": "Source_ig",
            "run": lambda wb, http_client, now_utc, args: scrape_instagram(
                wb, http_client, now_utc, args.days, args.max_per_profile, args.scrape_date
            ),
        },
        {
            "key": "linkedin",
            "label": "LinkedIn",
            "sheet": "Linkedin",
            "source_sheet": "Source_li",
            "run": lambda wb, http_client, now_utc, args: scrape_linkedin(
                wb, http_client, now_utc, args.days, args.max_per_profile, args.scrape_date
            ),
        },
        {
            "key": "tiktok",
            "label": "TikTok",
            "sheet": "TikTok",
            "source_sheet": "Source_tt",
            "run": lambda wb, http_client, now_utc, args: scrape_tiktok(
                wb, http_client, now_utc, args.days, args.max_per_profile, args.scrape_date
            ),
        },
        {
            "key": "facebook",
            "label": "Facebook",
            "sheet": "Facebook",
            "source_sheet": "Source_fb",
            "run": lambda wb, http_client, now_utc, args: scrape_facebook(
                wb, http_client, now_utc, args.scrape_date, days=args.days, max_per_profile=args.max_per_profile
            ),
        },
    ]


def main():
    parser = argparse.ArgumentParser(description="Multi-platform social scrape to Excel.")
    parser.add_argument("xlsx_path", help="Path to workbook")
    parser.add_argument("--days", type=int, default=7, help="Keep posts from last N days when date is available")
    parser.add_argument("--max-per-profile", type=int, default=5, help="Max posts per profile per run")
    parser.add_argument(
        "--skip-platforms",
        nargs="*",
        default=list(DEFAULT_SKIPPED_PLATFORMS),
        help="Optional list of platforms to skip. TikTok is skipped by default.",
    )
    parser.add_argument(
        "--rewrite-topics-only",
        action="store_true",
        help="Rewrite topic/topic_quality/topic_reason from raw_text; tries to backfill raw_text from post URLs; no new scraping rows",
    )
    args = parser.parse_args()

    wb = load_workbook(args.xlsx_path)
    platform_specs = get_platform_specs()
    skipped = {normalize_platform_key(name) for name in args.skip_platforms}
    skipped.discard("")
    enabled_specs = [spec for spec in platform_specs if spec["key"] not in skipped]
    enabled_sheets = [spec["sheet"] for spec in enabled_specs]

    required = set()
    for spec in enabled_specs:
        required.add(spec["sheet"])
        required.add(spec["source_sheet"])
    missing = [x for x in required if x not in wb.sheetnames]
    if missing:
        raise SystemExit(f"Missing sheets: {', '.join(missing)}")

    now_utc = datetime.now(timezone.utc)
    args.scrape_date = now_utc.astimezone().date().isoformat()
    http_client = HttpClient()
    ensure_scrape_headers_in_workbook(wb)
    cleanup = cleanup_workbook_layout(wb)
    topic_changes = normalize_topics_in_workbook(
        wb,
        http_client=http_client,
        backfill_from_urls=args.rewrite_topics_only,
        sheet_names=enabled_sheets,
    )

    if args.rewrite_topics_only:
        wb.save(args.xlsx_path)
        print("DONE")
        print(
            f"CLEANUP: fb_snapshots_removed={cleanup.get('Facebook_snapshots',0)} "
            f"tt_snapshots_removed={cleanup.get('TikTok_snapshots',0)}"
        )
        print(
            "TOPICS_REWRITTEN: "
            f"Facebook={topic_changes.get('Facebook',0)} "
            f"LinkedIn={topic_changes.get('Linkedin',0)} "
            f"YouTube={topic_changes.get('YouTube',0)} "
            f"TikTok={topic_changes.get('TikTok',0)} "
            f"Instagram={topic_changes.get('Instagram',0)}"
        )
        return

    summary = {}
    for spec in platform_specs:
        if spec["key"] in skipped:
            summary[spec["label"]] = {"profiles": 0, "added": 0, "skipped": 1}
            continue
        if spec["sheet"] not in wb.sheetnames or spec["source_sheet"] not in wb.sheetnames:
            summary[spec["label"]] = {"profiles": 0, "added": 0, "skipped": 1}
            continue
        summary[spec["label"]] = spec["run"](wb, http_client, now_utc, args)

    wb.save(args.xlsx_path)
    print("DONE")
    print(
        f"CLEANUP: fb_snapshots_removed={cleanup.get('Facebook_snapshots',0)} "
        f"tt_snapshots_removed={cleanup.get('TikTok_snapshots',0)}"
    )
    print(
        "TOPICS_REWRITTEN: "
        f"Facebook={topic_changes.get('Facebook',0)} "
        f"LinkedIn={topic_changes.get('Linkedin',0)} "
        f"YouTube={topic_changes.get('YouTube',0)} "
        f"TikTok={topic_changes.get('TikTok',0)} "
        f"Instagram={topic_changes.get('Instagram',0)}"
    )
    for platform in ["YouTube", "Instagram", "LinkedIn", "TikTok", "Facebook"]:
        s = summary.get(platform, {})
        suffix = " skipped=1" if s.get("skipped") else ""
        source_suffix = f" source={s.get('source')}" if s.get("source") else ""
        print(f"{platform}: profiles={s.get('profiles',0)} added={s.get('added',0)}{suffix}{source_suffix}")


if __name__ == "__main__":
    main()
